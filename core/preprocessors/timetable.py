from openpyxl import load_workbook
import re

import core.helpers.common_utils as common_utils
import core.helpers.course_utils as course_utils
from core.models.course import Course
import core.preprocessors.validators as validators


DAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday")


class Reader:
    def __init__(self, filename="core/source_files/latest.xlsx"):
        data = load_workbook(filename)
        sheets = data.sheetnames
        self.info = list(data[sheets[0]].values)
        self.periods = self.info[2]
        self.timings = self.info[3]
        self.content = self.info[4:-3]

    def get_venue(self, content):
        return content[1]

    def get_days_info(self):
        """
            Store information sorted by days.
            
            @returns:
                dictionary: key (day), value (day_information)
                
            @example:
                {
                    "Monday": [... information],
                 "Tuesday": [... information]
                 }
        """
        count = 0

        course_info = {k: [] for k in DAYS}
        # Day [0] -> Data Stored for Day 0 which is Monday
        # Day [1] -> Data Stored for Day 1 which is Tuesday

        for row in self.content:
            if row[0]:
                if row[0].strip().startswith(DAYS[count]):
                    count += 1
            course_info[DAYS[count - 1]].append(row)
            
        return course_info

    def get_courses_info(self):
        """
        Returns list of course objects
        """
        DAYS_INFO = self.get_days_info()
        courses_added = {}
        all_courses_info = []
        # TODO: Simplify this logic
        for day in DAYS:
            for column in range(len(DAYS_INFO[day])):
                for index, course_title in enumerate(DAYS_INFO[day][column]):
                    # A course title is of format: Digital Logic Design Lab (BCS-2M2)
                    if not validators.is_course_title(course_title):
                        continue

                    course_title = common_utils.remove_multiple_spaces(course_title)
                    print(course_title)
                    course_start_timing = course_utils.get_course_time(
                        course_title, index, self.timings, self.periods
                    )
                    is_lab_course = True if "lab" in course_title.lower() else False
                    course_end_timing = course_utils.generate_end_time(
                        course_start_timing, lab_course=is_lab_course
                    )
                    section = course_utils.get_section(course_title)

                    try:
                        course_title = course_title.rstrip(f"({section})").strip()
                        key = course_title + section

                        if not courses_added.get(key):
                            courses_added[key] = 1
                            if "lab" not in course_title.lower():
                                try:
                                    next_day = DAYS[DAYS.index(day) + 2][:3]
                                    days = f"{day[:3]}, {next_day}"
                                except IndexError:
                                    days = day
                                course_days = f"{days}"

                            else:
                                if ',' not in section and '&' not in section:
                                    section = section
                                else: 
                                    splitter = "," if "," in section else "&"
                                    section = section.split(splitter)[0].strip()[:-1]
                                course_days = day[:3]

                            semester = re.search(r"\d", section)

                            if semester:
                                semester = semester.group()
                            else:
                                semester = "unknown"

                            filtered_section = re.sub("[0-9]", "", section)

                            _course = Course(
                                name=course_title,
                                section="MCS" if not filtered_section else filtered_section,
                                start_time=course_start_timing,
                                end_time=course_end_timing,
                                room=self.get_venue(DAYS_INFO[day][column]),
                                day=course_days,
                                semester=semester,
                            )
                            all_courses_info.append(_course)

                    except Exception as e:
                        print(e)

        return sorted(all_courses_info, key=lambda x: x.name)

