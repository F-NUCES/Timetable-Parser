from openpyxl import load_workbook
import csv


def update_xlsx(src, dest):
    # Open an xlsx for reading
    wb = list(load_workbook(filename=dest).active.values)
    # Get the current Active Sheet
    wb[4:][1][2:][9] = "ABCDEFGH"
    print(wb[4:][1][2:])
    wb.save('newabc.xlsx')
    # You can also select a particular sheet
    # based on sheet name
    # ws = wb.get_sheet_by_name("Sheet1")
    # Open the csv file
    # with open(src) as fin:
        # read the csv
        # reader = csv.reader(fin)
        # enumerate the rows, so that you can
        # get the row index for the xlsx
        # for index, row in enumerate(reader):
        #     # Assuming space separated,
        #     # Split the row to cells (column)
        #     row = row[0].split()
        #     # Access the particular cell and assign
        #     # the value from the csv row
        #     ws.cell(row=index, column=7).value = row[2]
        #     ws.cell(row=index, column=8).value = row[3]
    # save the csv file
    # wb.save(dest)

update_xlsx("abc.xlsx", "source_files/new.xlsx")
